import csv
import copy
import fnmatch
import json
import logging
import posixpath
from concurrent.futures import ThreadPoolExecutor
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import httpx

from app.connectors.registry import get_connector
from app.db.database import db, decode, encode
from app.services.sql_safety import validate_source_query
from app.services.transforms import preview_transforms

logger = logging.getLogger(__name__)
executor = ThreadPoolExecutor(max_workers=4)


class RunStopped(RuntimeError):
    pass


def enqueue_run(pipeline_id: int, job_type: str = "manual", triggered_by: str = "system") -> int:
    with db() as conn:
        row = conn.execute(
            "INSERT INTO runs (pipeline_id, status) VALUES (?, 'queued') RETURNING id",
            (pipeline_id,),
        ).fetchone()
        run_id = int(dict(row)["id"])
    executor.submit(run_pipeline, run_id, job_type, triggered_by)
    return run_id


def preview(source_key: str, source_config: dict[str, Any], transforms: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = extract(source_key, source_config)
    return preview_transforms(rows[:25], prepare_runtime_transforms(transforms, source_key, source_config)).rows[:25]


def run_pipeline(run_id: int, job_type: str = "manual", triggered_by: str = "system") -> None:
    try:
        pipeline = _load_pipeline(run_id)
        if not _mark_running(run_id):
            _log(run_id, "INFO", "Run was stopped before it started")
            return
        _log(run_id, "INFO", f"Run started for pipeline {pipeline['name']}")
        _etl_audit_start(run_id, pipeline, job_type, triggered_by)
        _etl_audit_stage(run_id, "extract", source_path=_source_path(pipeline["source_key"], pipeline["source_config"]))
        rows = extract(pipeline["source_key"], pipeline["source_config"])
        _ensure_running(run_id)
        _update_counts(run_id, rows_read=len(rows))
        _etl_audit_stage(run_id, "transform", total_count=len(rows), source_path=_source_path(pipeline["source_key"], pipeline["source_config"], rows))
        _log(run_id, "INFO", f"Extracted {len(rows)} rows")
        result = preview_transforms(rows, prepare_runtime_transforms(pipeline["transforms"], pipeline["source_key"], pipeline["source_config"]))
        _ensure_running(run_id)
        rows = result.rows
        _etl_audit_stage(run_id, "reject", rejected_count=len(result.rejected_rows), failed_count=len(result.rejected_rows))
        rejected_path = save_rejected_records(pipeline["destination_key"], pipeline["destination_config"], result.rejected_rows)
        if rejected_path:
            _log(run_id, "WARNING", f"Rejected records saved: {rejected_path} ({len(result.rejected_rows)} rows)")
            _etl_audit_stage(run_id, "reject", error_file_path=rejected_path)
        for step_log in result.logs:
            _log(run_id, step_log.level, step_log.message)
            _transformation_log(run_id, step_log)
        for warning in result.warnings:
            _log(run_id, "WARNING", warning)
        _log(run_id, "INFO", f"Final output rows: {len(rows)}")
        _ensure_running(run_id)
        _etl_audit_stage(run_id, "load", target_path=_target_path(pipeline["destination_key"], pipeline["destination_config"]))
        written = load(pipeline["destination_key"], pipeline["destination_config"], rows)
        _ensure_running(run_id)
        _succeed(run_id, written)
        _log(run_id, "INFO", f"Run succeeded, wrote {written} rows")
        _etl_audit_finish(run_id, "succeeded", success_count=written)
    except RunStopped as exc:
        _log(run_id, "INFO", str(exc))
    except Exception as exc:
        logger.exception("Pipeline run failed")
        _fail(run_id, str(exc))
        _log(run_id, "ERROR", str(exc))
        _etl_audit_fail(run_id, str(exc))


def extract(source_key: str, config: dict[str, Any]) -> list[dict[str, Any]]:
    connector = get_connector(source_key)
    if connector.type != "source":
        raise ValueError(f"{source_key} is not a source")
    if source_key == "sample_crm":
        limit = int(config.get("limit", 25))
        return [
            {"id": idx, "name": f"Customer {idx}", "tier": "enterprise" if idx % 3 == 0 else "growth", "mrr": idx * 125}
            for idx in range(1, limit + 1)
        ]
    if source_key == "http_json":
        url = config["url"]
        response = httpx.get(url, timeout=30)
        response.raise_for_status()
        payload = response.json()
        if isinstance(payload, list):
            return [dict(item) for item in payload]
        if isinstance(payload, dict):
            records = payload.get("data") or payload.get("items") or payload.get("records")
            if isinstance(records, list):
                return [dict(item) for item in records]
            return [payload]
    if source_key == "csv_file":
        path = Path(config["path"])
        with path.open(newline="", encoding="utf-8") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if source_key == "postgres_source":
        return _extract_postgres(config)
    if source_key == "sftp_source":
        return _extract_sftp(config)
    raise ValueError(f"Unsupported source {source_key}")


def _extract_postgres(config: dict[str, Any]) -> list[dict[str, Any]]:
    try:
        import psycopg
        from psycopg.rows import dict_row
    except ImportError as exc:
        raise RuntimeError("psycopg is required for PostgreSQL extraction. Install backend requirements.") from exc

    query = config.get("query")
    if query:
        query = validate_source_query(str(query))
    else:
        schema = "".join(ch for ch in config.get("schema", "public") if ch.isalnum() or ch == "_")
        table = "".join(ch for ch in config["table"] if ch.isalnum() or ch == "_")
        query = f'SELECT * FROM "{schema}"."{table}" LIMIT 1000'
    with psycopg.connect(
        host=config["host"],
        port=int(config.get("port", 5432)),
        dbname=config["database"],
        user=config["username"],
        password=config.get("password", ""),
        connect_timeout=10,
        row_factory=dict_row,
    ) as conn:
        return [dict(row) for row in conn.execute(query).fetchall()]


def _extract_sftp(config: dict[str, Any]) -> list[dict[str, Any]]:
    try:
        import io

        import paramiko
    except ImportError as exc:
        raise RuntimeError("paramiko is required for SFTP extraction. Install backend requirements.") from exc

    password = config.get("password") or None
    private_key = config.get("private_key") or None
    pkey = paramiko.RSAKey.from_private_key(io.StringIO(private_key)) if private_key else None
    transport = paramiko.Transport((config["host"], int(config.get("port", 22))))
    try:
        transport.connect(username=config["username"], password=password, pkey=pkey)
        client = paramiko.SFTPClient.from_transport(transport)
        rows: list[dict[str, Any]] = []
        for remote_path in _sftp_read_paths(client, config):
            with client.open(remote_path, "r") as handle:
                content = handle.read()
            if config.get("format") == "xlsx" or remote_path.endswith(".xlsx"):
                file_rows = _rows_from_xlsx(content if isinstance(content, bytes) else content.encode("utf-8"), config.get("sheet_name"), config.get("file_password"))
            else:
                if isinstance(content, bytes):
                    content = content.decode("utf-8")
                file_rows = [dict(row) for row in csv.DictReader(content.splitlines())]
            for row in file_rows:
                row.setdefault("_source_file", remote_path)
            rows.extend(file_rows)
        return rows
    finally:
        transport.close()


def save_rejected_records(destination_key: str, config: dict[str, Any], rows: list[dict[str, Any]]) -> str | None:
    if not rows:
        return None
    if destination_key == "sftp_destination":
        remote_path = _rejected_write_path(config, _sftp_write_path(config))
        _write_sftp_rows(config, remote_path, rows)
        return remote_path
    if destination_key in {"csv_output", "jsonl_file"}:
        output_path = str(config.get("path", "data/output.csv"))
        rejected_path = Path(_rejected_write_path(config, output_path))
        rejected_path.parent.mkdir(parents=True, exist_ok=True)
        rejected_path.write_bytes(_rows_payload(str(rejected_path), rows, config))
        return str(rejected_path)
    return None


def load(destination_key: str, config: dict[str, Any], rows: list[dict[str, Any]]) -> int:
    connector = get_connector(destination_key)
    if connector.type != "destination":
        raise ValueError(f"{destination_key} is not a destination")
    if destination_key == "jsonl_file":
        path = Path(config.get("path", "data/output.jsonl"))
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps(row) + "\n")
        return len(rows)
    if destination_key == "csv_output":
        path = Path(config.get("path", "data/output.csv"))
        path.parent.mkdir(parents=True, exist_ok=True)
        if not rows:
            return 0
        exists = path.exists()
        with path.open("a", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=_row_columns(rows))
            if not exists:
                writer.writeheader()
            writer.writerows(rows)
        return len(rows)
    if destination_key == "postgres_destination":
        return _load_postgres(config, rows)
    if destination_key == "sftp_destination":
        return _load_sftp(config, rows)
    raise ValueError(f"Unsupported destination {destination_key}")


def _load_postgres(config: dict[str, Any], rows: list[dict[str, Any]]) -> int:
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for PostgreSQL destination. Install backend requirements.") from exc

    schema = "".join(ch for ch in config.get("schema", "public") if ch.isalnum() or ch == "_")
    table = "".join(ch for ch in config["table"] if ch.isalnum() or ch == "_")
    columns = [str(key) for key in rows[0].keys()] if rows else []
    sql = _postgres_write_sql(schema, table, columns, config) if rows else ""
    with psycopg.connect(
        host=config["host"],
        port=int(config.get("port", 5432)),
        dbname=config["database"],
        user=config["username"],
        password=config.get("password", ""),
        connect_timeout=10,
    ) as conn:
        with conn.cursor() as cursor:
            if config.get("mode") == "truncate_insert":
                cursor.execute(f"TRUNCATE TABLE {_quote_identifier(schema)}.{_quote_identifier(table)}")
            if rows:
                cursor.executemany(sql, [tuple(row.get(col) for col in columns) for row in rows])
    return len(rows)


def _postgres_write_sql(schema: str, table: str, columns: list[str], config: dict[str, Any]) -> str:
    quoted_columns = ", ".join(_quote_identifier(col) for col in columns)
    placeholders = ", ".join(["%s"] * len(columns))
    sql = f"INSERT INTO {_quote_identifier(schema)}.{_quote_identifier(table)} ({quoted_columns}) VALUES ({placeholders})"
    if config.get("mode") in {"truncate_insert", "truncate"}:
        return sql
    if config.get("mode") != "upsert":
        return sql
    primary_key = str(config.get("primary_key") or "").strip()
    if not primary_key:
        raise ValueError("PostgreSQL upsert needs primary_key")
    if primary_key not in columns:
        raise ValueError(f"PostgreSQL upsert primary key {primary_key} is not in output rows")
    update_columns = [col for col in columns if col != primary_key]
    if not update_columns:
        return f"{sql} ON CONFLICT ({_quote_identifier(primary_key)}) DO NOTHING"
    assignments = ", ".join(f"{_quote_identifier(col)}=EXCLUDED.{_quote_identifier(col)}" for col in update_columns)
    return f"{sql} ON CONFLICT ({_quote_identifier(primary_key)}) DO UPDATE SET {assignments}"


def _quote_identifier(value: str) -> str:
    if not value:
        raise ValueError("PostgreSQL identifier cannot be empty")
    return '"' + value.replace('"', '""') + '"'


def _load_sftp(config: dict[str, Any], rows: list[dict[str, Any]]) -> int:
    try:
        import io

        import paramiko
    except ImportError as exc:
        raise RuntimeError("paramiko is required for SFTP destination. Install backend requirements.") from exc
    remote_path = _sftp_write_path(config)
    if config.get("format") == "xlsx" or remote_path.endswith(".xlsx"):
        payload: str | bytes = _xlsx_from_rows(rows, config)
    else:
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=_row_columns(rows))
            writer.writeheader()
            writer.writerows(rows)
        payload = output.getvalue()
    _write_sftp_payload(config, remote_path, payload)
    return len(rows)


def _write_sftp_rows(config: dict[str, Any], remote_path: str, rows: list[dict[str, Any]]) -> None:
    _write_sftp_payload(config, remote_path, _rows_payload(remote_path, rows, config))


def _write_sftp_payload(config: dict[str, Any], remote_path: str, payload: str | bytes) -> None:
    try:
        import io

        import paramiko
    except ImportError as exc:
        raise RuntimeError("paramiko is required for SFTP destination. Install backend requirements.") from exc
    password = config.get("password") or None
    private_key = config.get("private_key") or None
    pkey = paramiko.RSAKey.from_private_key(io.StringIO(private_key)) if private_key else None
    transport = paramiko.Transport((config["host"], int(config.get("port", 22))))
    try:
        transport.connect(username=config["username"], password=password, pkey=pkey)
        client = paramiko.SFTPClient.from_transport(transport)
        if _truthy(config.get("auto_create_folders"), default=True):
            _ensure_sftp_directory(client, remote_path)
        with client.open(remote_path, "w") as handle:
            handle.write(payload)
    finally:
        transport.close()


def _ensure_sftp_directory(client: Any, remote_path: str) -> None:
    directory = posixpath.dirname(remote_path)
    if not directory or directory in {".", "/"}:
        return
    current = "/" if directory.startswith("/") else ""
    for part in [item for item in directory.split("/") if item]:
        current = posixpath.join(current, part) if current else part
        try:
            client.stat(current)
        except OSError:
            client.mkdir(current)


def _truthy(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _row_columns(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            column = str(key)
            if column not in seen:
                seen.add(column)
                columns.append(column)
    return columns


def _rows_payload(path: str, rows: list[dict[str, Any]], config: dict[str, Any] | None = None) -> bytes:
    if path.endswith(".xlsx"):
        payload = _xlsx_from_rows(rows, config)
        return payload if isinstance(payload, bytes) else payload.encode("utf-8")
    if path.endswith(".jsonl"):
        return "".join(json.dumps(row, default=str) + "\n" for row in rows).encode("utf-8")
    import io

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=_row_columns(rows))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


def prepare_runtime_transforms(transforms: list[dict[str, Any]], base_source_key: str | None = None, base_source_config: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    prepared = copy.deepcopy(transforms)
    for step in prepared:
        step_type = step.get("step_type") or step.get("type")
        params = step.setdefault("parameters", {})
        if step_type == "join" and not params.get("right_rows"):
            if params.get("right_source_mode") == "same_connection":
                if not base_source_key or not base_source_config:
                    raise ValueError("Join source connection is not available")
                right_config = _same_connection_join_config(base_source_key, base_source_config, params.get("right_source_config") or {})
                params["right_rows"] = extract(base_source_key, right_config)
            elif params.get("right_source_id"):
                resource = _resource_for_join(int(params["right_source_id"]))
                params["right_rows"] = extract(resource["connector_key"], resource["config"])
    return prepared


def _resource_for_join(resource_id: int) -> dict[str, Any]:
    with db() as conn:
        row = conn.execute("SELECT connector_key, config FROM resources WHERE id=? AND type='source'", (resource_id,)).fetchone()
    if row is None:
        raise ValueError(f"Join source not found: {resource_id}")
    data = dict(row)
    data["config"] = decode(data["config"])
    return data


def _same_connection_join_config(base_source_key: str, base_source_config: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(overrides, dict):
        raise ValueError("Join source override config is invalid")
    if base_source_key == "postgres_source":
        next_config = dict(base_source_config)
        for key in ("schema", "table", "query"):
            value = overrides.get(key)
            if value not in (None, ""):
                next_config[key] = value
        if next_config.get("query"):
            next_config["table"] = overrides.get("table") or next_config.get("table", "")
        elif not next_config.get("table"):
            raise ValueError("Join source needs table or query")
        return next_config
    if base_source_key == "sftp_source":
        next_config = dict(base_source_config)
        base_remote_path = str(base_source_config.get("remote_path") or "")
        remote_path_override = overrides.get("remote_path")
        if remote_path_override not in (None, ""):
            next_config["remote_path"] = _resolve_sftp_join_path(base_remote_path, str(remote_path_override))
        next_config.pop("path_pattern", None)
        format_override = overrides.get("format")
        if format_override not in (None, ""):
            next_config["format"] = format_override
        if not next_config.get("remote_path"):
            raise ValueError("Join source needs remote path")
        return next_config
    return {**base_source_config, **overrides}


def _resolve_sftp_join_path(base_remote_path: str, candidate: str) -> str:
    if not candidate or candidate.startswith("/"):
        return candidate
    base = base_remote_path.rstrip("/")
    if not base:
        return candidate
    return posixpath.join(base, candidate)


def _sftp_read_paths(client, config: dict[str, Any]) -> list[str]:
    remote_path = str(config.get("remote_path") or "").strip()
    if remote_path:
        return [_format_path_pattern(remote_path)]
    pattern = str(config.get("path_pattern") or "").strip()
    if not pattern:
        raise ValueError("SFTP source needs remote path")
    pattern = _format_path_pattern(pattern)
    if not any(char in pattern for char in "*?[]"):
        return [pattern]
    directory = posixpath.dirname(pattern) or "."
    filename_pattern = posixpath.basename(pattern)
    return [posixpath.join(directory, item) for item in client.listdir(directory) if fnmatch.fnmatch(item, filename_pattern)]


def _sftp_write_path(config: dict[str, Any]) -> str:
    remote_path = str(config.get("remote_path") or "").strip()
    pattern = str(config.get("output_path_pattern") or "").strip()
    if remote_path:
        remote_path = _format_path_pattern(remote_path)
    elif pattern:
        remote_path = _format_path_pattern(pattern)
    if not remote_path:
        raise ValueError("SFTP destination needs output path")
    if remote_path.endswith("/"):
        default_name = "output.xlsx" if config.get("format") == "xlsx" else "output.csv"
        return posixpath.join(remote_path, default_name)
    return remote_path


def _rejected_output_path(output_path: str) -> str:
    timestamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S")
    if "." in posixpath.basename(output_path):
        stem, ext = posixpath.splitext(output_path)
        return f"{stem}_rejected_{timestamp}{ext}"
    return posixpath.join(output_path, f"rejected_{timestamp}.csv")


def _rejected_write_path(config: dict[str, Any], output_path: str) -> str:
    explicit = str(config.get("rejected_path") or config.get("error_path") or "").strip()
    pattern = str(config.get("rejected_path_pattern") or config.get("error_path_pattern") or "").strip()
    if explicit:
        return _format_path_pattern(explicit)
    if pattern:
        return _format_path_pattern(pattern)
    return _rejected_output_path(output_path)


def _format_path_pattern(pattern: str) -> str:
    now = datetime.now(UTC)
    values = {
        "YYYY": now.strftime("%Y"),
        "YY": now.strftime("%y"),
        "MM": now.strftime("%m"),
        "DD": now.strftime("%d"),
        "hh": now.strftime("%H"),
        "mm": now.strftime("%M"),
        "ss": now.strftime("%S"),
        "timestamp": now.strftime("%Y%m%d%H%M%S"),
    }
    return pattern.format(**values)


def _rows_from_xlsx(content: bytes, sheet_name: Any = None, password: Any = None) -> list[dict[str, Any]]:
    import io

    from openpyxl import load_workbook

    content = _decrypt_xlsx_if_needed(content, password)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    _validate_single_visible_sheet(workbook)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [{headers[index]: value for index, value in enumerate(row)} for row in rows[1:]]


def _decrypt_xlsx_if_needed(content: bytes, password: Any = None) -> bytes:
    if not password:
        return content
    try:
        import io
        import msoffcrypto
    except ImportError as exc:
        raise RuntimeError("Password-protected XLSX input requires msoffcrypto-tool. Install backend requirement and set file_password.") from exc
    decrypted = io.BytesIO()
    office_file = msoffcrypto.OfficeFile(io.BytesIO(content))
    office_file.load_key(password=str(password))
    office_file.decrypt(decrypted)
    return decrypted.getvalue()


def _validate_single_visible_sheet(workbook: Any) -> None:
    sheet_names = list(workbook.sheetnames)
    hidden = [sheet.title for sheet in workbook.worksheets if getattr(sheet, "sheet_state", "visible") != "visible"]
    if hidden:
        raise ValueError(f"XLSX input contains hidden sheets: {', '.join(hidden)}")
    if len(sheet_names) != 1:
        raise ValueError(f"XLSX input must contain exactly one sheet; found {len(sheet_names)}")


def _xlsx_from_rows(rows: list[dict[str, Any]], config: dict[str, Any] | None = None) -> bytes:
    import io

    from openpyxl import Workbook

    config = config or {}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _xlsx_title(str(config.get("xlsx_data_sheet") or "Data"))
    if rows:
        headers = _row_columns(rows)
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    cell.number_format = "@"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_title(value: str) -> str:
    cleaned = "".join(char for char in value if char not in "[]:*?/\\")[:31].strip()
    return cleaned or "Sheet1"


def _load_pipeline(run_id: int) -> dict[str, Any]:
    with db() as conn:
        row = conn.execute(
            """
            SELECT p.* FROM pipelines p
            JOIN runs r ON r.pipeline_id = p.id
            WHERE r.id = ?
            """,
            (run_id,),
        ).fetchone()
    if row is None:
        raise ValueError(f"Run not found: {run_id}")
    data = dict(row)
    data["source_config"] = decode(data["source_config"])
    data["destination_config"] = decode(data["destination_config"])
    data["transforms"] = _runtime_pipeline_transforms(data)
    return data


def _runtime_pipeline_transforms(pipeline: dict[str, Any]) -> list[dict[str, Any]]:
    fallback = decode(pipeline["transforms"])
    transformation_id = pipeline.get("transformation_id") or _matching_transformation_id(fallback)
    if not transformation_id:
        return fallback
    version = pipeline.get("transformation_version")
    with db() as conn:
        if version:
            row = conn.execute(
                """
                SELECT snapshot_data
                FROM transformation_versions
                WHERE transformation_id=? AND version_no=?
                ORDER BY id DESC
                LIMIT 1
                """,
                (transformation_id, version),
            ).fetchone()
        else:
            row = conn.execute(
                """
                SELECT snapshot_data
                FROM transformation_versions
                WHERE transformation_id=?
                ORDER BY version_no DESC, id DESC
                LIMIT 1
                """,
                (transformation_id,),
            ).fetchone()
    if row is None:
        return fallback
    snapshot = decode(dict(row)["snapshot_data"])
    steps = snapshot.get("steps") if isinstance(snapshot, dict) else None
    return steps if isinstance(steps, list) else fallback


def _matching_transformation_id(steps: list[dict[str, Any]]) -> int | None:
    with db() as conn:
        version_rows = conn.execute(
            """
            SELECT transformation_id, snapshot_data
            FROM transformation_versions
            ORDER BY version_no DESC, id DESC
            """
        ).fetchall()
        transformation_rows = conn.execute(
            """
            SELECT id, steps
            FROM transformations
            ORDER BY id DESC
            """
        ).fetchall()
    for row in version_rows:
        data = dict(row)
        snapshot = decode(data["snapshot_data"])
        if isinstance(snapshot, dict) and snapshot.get("steps") == steps:
            return int(data["transformation_id"])
    for row in transformation_rows:
        data = dict(row)
        if decode(data["steps"]) == steps:
            return int(data["id"])
    return None


def _mark_running(run_id: int) -> bool:
    with db() as conn:
        cursor = conn.execute("UPDATE runs SET status='running', started_at=CURRENT_TIMESTAMP WHERE id=? AND status='queued'", (run_id,))
        return cursor.rowcount == 1


def _ensure_running(run_id: int) -> None:
    with db() as conn:
        row = conn.execute("SELECT status, error FROM runs WHERE id=?", (run_id,)).fetchone()
    if row is None:
        raise RunStopped(f"Run {run_id} no longer exists")
    data = dict(row)
    if data["status"] != "running":
        reason = data.get("error") or data["status"]
        raise RunStopped(f"Run stopped: {reason}")


def _update_counts(run_id: int, rows_read: int) -> None:
    with db() as conn:
        conn.execute("UPDATE runs SET rows_read=? WHERE id=?", (rows_read, run_id))


def _succeed(run_id: int, rows_written: int) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE runs SET status='succeeded', rows_written=?, finished_at=CURRENT_TIMESTAMP WHERE id=? AND status='running'",
            (rows_written, run_id),
        )


def _fail(run_id: int, error: str) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE runs SET status='failed', error=?, finished_at=CURRENT_TIMESTAMP WHERE id=? AND status IN ('queued', 'running')",
            (error, run_id),
        )


def _etl_audit_start(run_id: int, pipeline: dict[str, Any], job_type: str, triggered_by: str) -> None:
    now = datetime.now(UTC).isoformat()
    with db() as conn:
        conn.execute(
            """
            INSERT INTO etl_audit_log
            (run_id, pipeline_name, job_type, start_time, status, current_stage, source_path, target_path, triggered_by, created_date, last_modified_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                pipeline.get("name"),
                job_type,
                now,
                "running",
                "started",
                _source_path(pipeline["source_key"], pipeline["source_config"]),
                _target_path(pipeline["destination_key"], pipeline["destination_config"]),
                triggered_by,
                now,
                now,
            ),
        )


def _etl_audit_stage(run_id: int, stage: str, **fields: object) -> None:
    _etl_audit_update(run_id, current_stage=stage, **fields)


def _etl_audit_finish(run_id: int, status: str, **fields: object) -> None:
    end_time = datetime.now(UTC).isoformat()
    start_time = _etl_audit_start_time(run_id)
    duration = _duration_seconds(start_time, end_time)
    _etl_audit_update(run_id, status=status, current_stage="completed", end_time=end_time, duration_seconds=duration, **fields)


def _etl_audit_fail(run_id: int, error: str) -> None:
    stage = _etl_audit_current_stage(run_id) or "unknown"
    end_time = datetime.now(UTC).isoformat()
    start_time = _etl_audit_start_time(run_id)
    duration = _duration_seconds(start_time, end_time)
    _etl_audit_update(run_id, status="failed", current_stage="failed", failed_stage=stage, end_time=end_time, duration_seconds=duration, error_message=error)


def mark_run_stopped_audit(run_id: int, triggered_by: str) -> None:
    stage = _etl_audit_current_stage(run_id) or "stopped"
    end_time = datetime.now(UTC).isoformat()
    start_time = _etl_audit_start_time(run_id)
    duration = _duration_seconds(start_time, end_time)
    _etl_audit_update(
        run_id,
        status="stopped",
        current_stage="stopped",
        failed_stage=stage,
        end_time=end_time,
        duration_seconds=duration,
        error_message=f"Stopped by {triggered_by}",
    )


def _etl_audit_update(run_id: int, **fields: object) -> None:
    allowed = {
        "status", "current_stage", "failed_stage", "source_path", "target_path", "total_count", "success_count",
        "failed_count", "rejected_count", "error_message", "error_file_path", "end_time", "duration_seconds",
    }
    updates = {key: value for key, value in fields.items() if key in allowed}
    if not updates:
        return
    updates["last_modified_date"] = datetime.now(UTC).isoformat()
    assignments = ", ".join(f"{key}=?" for key in updates)
    values = tuple(updates.values()) + (run_id,)
    with db() as conn:
        conn.execute(f"UPDATE etl_audit_log SET {assignments} WHERE run_id=?", values)


def _etl_audit_start_time(run_id: int) -> str | None:
    with db() as conn:
        row = conn.execute("SELECT start_time FROM etl_audit_log WHERE run_id=? ORDER BY id DESC LIMIT 1", (run_id,)).fetchone()
    return str(dict(row).get("start_time")) if row and dict(row).get("start_time") else None


def _etl_audit_current_stage(run_id: int) -> str | None:
    with db() as conn:
        row = conn.execute("SELECT current_stage FROM etl_audit_log WHERE run_id=? ORDER BY id DESC LIMIT 1", (run_id,)).fetchone()
    return str(dict(row).get("current_stage")) if row and dict(row).get("current_stage") else None


def _duration_seconds(start_time: str | None, end_time: str) -> float | None:
    if not start_time:
        return None
    try:
        return round((datetime.fromisoformat(end_time) - datetime.fromisoformat(start_time)).total_seconds(), 3)
    except ValueError:
        return None


def _source_path(source_key: str, config: dict[str, Any], rows: list[dict[str, Any]] | None = None) -> str:
    if rows:
        files = sorted({str(row.get("_source_file")) for row in rows if row.get("_source_file")})
        if files:
            return ",".join(files)
    if source_key == "sftp_source":
        return _format_path_pattern(str(config.get("remote_path") or config.get("path_pattern") or ""))
    if source_key == "csv_file":
        return str(config.get("path") or "")
    if source_key == "postgres_source":
        if config.get("query"):
            return "query"
        return ".".join(str(part) for part in (config.get("schema", "public"), config.get("table", "")) if part)
    return source_key


def _target_path(destination_key: str, config: dict[str, Any]) -> str:
    if destination_key == "sftp_destination":
        try:
            return _sftp_write_path(config)
        except ValueError:
            return str(config.get("remote_path") or config.get("output_path_pattern") or "")
    if destination_key in {"csv_output", "jsonl_file"}:
        return str(config.get("path") or "")
    if destination_key == "postgres_destination":
        return ".".join(str(part) for part in (config.get("schema", "public"), config.get("table", "")) if part)
    return destination_key


def _log(run_id: int, level: str, message: str) -> None:
    logger.log(getattr(logging, level, logging.INFO), message)
    with db() as conn:
        conn.execute("INSERT INTO run_logs (run_id, level, message) VALUES (?, ?, ?)", (run_id, level, message))


def _transformation_log(run_id: int, step_log) -> None:
    with db() as conn:
        conn.execute(
            """
            INSERT INTO transformation_run_logs
            (run_id, step_id, status, message, records_before, records_after, duration_ms)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                step_log.step_id,
                step_log.level,
                step_log.message,
                step_log.records_before,
                step_log.records_after,
                step_log.duration_ms,
            ),
        )
