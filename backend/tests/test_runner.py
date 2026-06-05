from app.services.runner import _postgres_write_sql, _same_connection_join_config
from app.services.sql_safety import validate_source_query


def test_same_connection_postgres_join_config():
    result = _same_connection_join_config(
        "postgres_source",
        {
            "host": "db.local",
            "port": 5432,
            "database": "warehouse",
            "username": "etl",
            "password": "secret",
            "schema": "public",
            "table": "customers",
        },
        {"schema": "analytics", "table": "loans"},
    )

    assert result["host"] == "db.local"
    assert result["database"] == "warehouse"
    assert result["schema"] == "analytics"
    assert result["table"] == "loans"


def test_same_connection_sftp_join_config():
    result = _same_connection_join_config(
        "sftp_source",
        {
            "host": "sftp.local",
            "port": 22,
            "username": "etl",
            "password": "secret",
            "remote_path": "/in",
            "format": "csv",
        },
        {"remote_path": "loans.csv", "format": "xlsx"},
    )

    assert result["host"] == "sftp.local"
    assert result["remote_path"] == "/in/loans.csv"
    assert result["format"] == "xlsx"
    assert "operation" not in result


def test_postgres_write_sql_supports_append_and_upsert():
    append_sql = _postgres_write_sql("public", "customers", ["id", "name"], {"mode": "append"})
    upsert_sql = _postgres_write_sql("public", "customers", ["id", "name"], {"mode": "upsert", "primary_key": "id"})

    assert append_sql == 'INSERT INTO "public"."customers" ("id", "name") VALUES (%s, %s)'
    assert upsert_sql == (
        'INSERT INTO "public"."customers" ("id", "name") VALUES (%s, %s) '
        'ON CONFLICT ("id") DO UPDATE SET "name"=EXCLUDED."name"'
    )


def test_postgres_upsert_requires_primary_key_in_rows():
    try:
        _postgres_write_sql("public", "customers", ["id", "name"], {"mode": "upsert", "primary_key": "customer_id"})
    except ValueError as exc:
        assert "primary key customer_id is not in output rows" in str(exc)
    else:
        raise AssertionError("Expected missing primary key to fail")


def test_validate_source_query_accepts_single_select():
    assert validate_source_query("select * from customers;") == "select * from customers"


def test_validate_source_query_blocks_mutation_keywords():
    try:
        validate_source_query("delete from customers")
    except ValueError as exc:
        assert "SELECT or WITH" in str(exc)
    else:
        raise AssertionError("Expected mutating query to fail")


def test_validate_source_query_blocks_multiple_statements():
    try:
        validate_source_query("select * from customers; drop table customers")
    except ValueError as exc:
        assert "single read-only statement" in str(exc)
    else:
        raise AssertionError("Expected multiple statements to fail")


def test_rows_payload_csv_uses_union_of_all_columns():
    from app.services.runner import _rows_payload

    payload = _rows_payload("/out/final.csv", [{"a": "1"}, {"a": "2", "b": "late"}])

    assert payload.decode("utf-8").splitlines()[0] == "a,b"


def test_xlsx_from_rows_uses_union_of_all_columns():
    import io

    from openpyxl import load_workbook

    from app.services.runner import _xlsx_from_rows

    payload = _xlsx_from_rows([{"a": "1"}, {"a": "2", "b": "late"}])
    workbook = load_workbook(io.BytesIO(payload), read_only=True)
    headers = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]

    assert headers == ["a", "b"]


def test_rejected_rows_payload_uses_jsonl_for_jsonl_path():
    from app.services.runner import _rows_payload

    payload = _rows_payload("/out/final_rejected.jsonl", [{"id": "2", "_rejected_reason": "Invalid float value"}])

    assert payload.decode("utf-8").splitlines() == ['{"id": "2", "_rejected_reason": "Invalid float value"}']


def test_rejected_output_path_adds_datetime_before_extension(monkeypatch):
    from datetime import UTC, datetime

    from app.services import runner

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 5, 25, 6, 30, 5, tzinfo=UTC)

    monkeypatch.setattr(runner, "datetime", FixedDateTime)

    assert runner._rejected_output_path("/out/final.csv") == "/out/final_rejected_20260525063005.csv"



def test_sftp_read_paths_formats_current_date_pattern(monkeypatch):
    from datetime import UTC, datetime

    from app.services import runner

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 5, 7, 8, 9, tzinfo=UTC)

    class Client:
        def listdir(self, directory):
            assert directory == "/in"
            return ["loan_20260605.csv", "loan_20260604.csv"]

    monkeypatch.setattr(runner, "datetime", FixedDateTime)

    assert runner._sftp_read_paths(Client(), {"path_pattern": "/in/loan_{YYYY}{MM}{DD}.csv"}) == ["/in/loan_20260605.csv"]


def test_rejected_write_path_uses_configured_error_pattern(monkeypatch):
    from datetime import UTC, datetime

    from app.services import runner

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 5, 7, 8, 9, tzinfo=UTC)

    monkeypatch.setattr(runner, "datetime", FixedDateTime)

    path = runner._rejected_write_path({"rejected_path_pattern": "/err/rejected_{YYYY}{MM}{DD}_{timestamp}.jsonl"}, "/out/final.csv")

    assert path == "/err/rejected_20260605_20260605070809.jsonl"
