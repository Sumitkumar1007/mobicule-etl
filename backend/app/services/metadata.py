import csv
import io
import posixpath
import stat
from pathlib import Path
from typing import Any

import httpx

from app.services.runner import extract
from app.services.sql_safety import validate_source_query


def source_columns(source_key: str, config: dict[str, Any]) -> list[str]:
    if source_key == "sample_crm":
        return ["id", "name", "tier", "mrr"]
    if source_key == "csv_file":
        path = Path(config["path"])
        with path.open(newline="", encoding="utf-8") as handle:
            reader = csv.reader(handle)
            return next(reader, [])
    if source_key == "http_json":
        rows = extract(source_key, config)[:1]
        return list(rows[0].keys()) if rows else []
    if source_key == "postgres_source":
        return _postgres_columns(config)
    if source_key == "sftp_source":
        return _sftp_columns(config)
    rows = extract(source_key, config)[:1]
    return list(rows[0].keys()) if rows else []


def source_options(source_key: str, config: dict[str, Any]) -> dict[str, list[str] | str]:
    if source_key == "postgres_source":
        return {"tables": _postgres_tables(config)}
    if source_key == "sftp_source":
        return _sftp_entries(config)
    return {}


def _postgres_columns(config: dict[str, Any]) -> list[str]:
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for PostgreSQL metadata. Install backend requirements.") from exc

    table = config.get("table")
    schema = config.get("schema", "public")
    query = config.get("query")
    with psycopg.connect(
        host=config["host"],
        port=int(config.get("port", 5432)),
        dbname=config["database"],
        user=config["username"],
        password=config.get("password", ""),
        connect_timeout=10,
    ) as conn:
        if table:
            rows = conn.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                ORDER BY ordinal_position
                """,
                (schema, table),
            ).fetchall()
            return [row[0] for row in rows]
        if query:
            try:
                safe_query = validate_source_query(str(query))
                cursor = conn.execute(f"SELECT * FROM ({safe_query}) AS preview_source LIMIT 0")
            except Exception as exc:
                raise ValueError(
                    "PostgreSQL column fetch failed. Fix the SQL query, or clear Query and use Schema/Table fields. "
                    f"Database error: {exc}"
                ) from exc
            return [column.name for column in cursor.description or []]
    return []


def _postgres_tables(config: dict[str, Any]) -> list[str]:
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for PostgreSQL metadata. Install backend requirements.") from exc

    schema = config.get("schema", "public")
    with psycopg.connect(
        host=config["host"],
        port=int(config.get("port", 5432)),
        dbname=config["database"],
        user=config["username"],
        password=config.get("password", ""),
        connect_timeout=10,
    ) as conn:
        rows = conn.execute(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = %s AND table_type IN ('BASE TABLE', 'VIEW')
            ORDER BY table_name
            """,
            (schema,),
        ).fetchall()
    return [row[0] for row in rows]


def _sftp_columns(config: dict[str, Any]) -> list[str]:
    try:
        import paramiko
    except ImportError as exc:
        raise RuntimeError("paramiko is required for SFTP metadata. Install backend requirements.") from exc

    host = config["host"]
    port = int(config.get("port", 22))
    username = config["username"]
    remote_path = config["remote_path"]
    password = config.get("password") or None
    private_key = config.get("private_key") or None
    pkey = paramiko.RSAKey.from_private_key(io.StringIO(private_key)) if private_key else None

    transport = paramiko.Transport((host, port))
    try:
        transport.connect(username=username, password=password, pkey=pkey)
        client = paramiko.SFTPClient.from_transport(transport)
        with client.open(remote_path, "r") as handle:
            if config.get("format") == "xlsx" or remote_path.endswith(".xlsx"):
                return _xlsx_columns(handle.read())
            first_line = handle.readline()
        return next(csv.reader([first_line]), [])
    finally:
        transport.close()


def _sftp_entries(config: dict[str, Any]) -> dict[str, list[str] | str]:
    try:
        import paramiko
    except ImportError as exc:
        raise RuntimeError("paramiko is required for SFTP metadata. Install backend requirements.") from exc

    host = config["host"]
    port = int(config.get("port", 22))
    username = config["username"]
    password = config.get("password") or None
    private_key = config.get("private_key") or None
    pkey = paramiko.RSAKey.from_private_key(io.StringIO(private_key)) if private_key else None
    seed = _sftp_seed_path(config)

    transport = paramiko.Transport((host, port))
    try:
        transport.connect(username=username, password=password, pkey=pkey)
        client = paramiko.SFTPClient.from_transport(transport)
        directory = _sftp_option_directory(client, seed)
        dirs: list[str] = []
        files: list[str] = []
        for item in client.listdir_attr(directory):
            path = posixpath.join(directory, item.filename)
            if stat.S_ISDIR(item.st_mode):
                dirs.append(path)
            else:
                files.append(path)
        return {"current_path": directory, "dirs": sorted(dirs), "paths": sorted(files)}
    finally:
        transport.close()


def _sftp_seed_path(config: dict[str, Any]) -> str:
    for key in ("remote_path", "path_pattern", "output_path_pattern"):
        value = str(config.get(key) or "").strip()
        if value and value not in {"path_pattern", "output_path_pattern"}:
            return value
    return "."


def _sftp_option_directory(client, seed: str) -> str:
    if any(char in seed for char in "*?[]"):
        return posixpath.dirname(seed) or "."
    try:
        mode = client.stat(seed).st_mode
        if stat.S_ISDIR(mode):
            return seed
    except Exception:
        pass
    return posixpath.dirname(seed) or "."


def _xlsx_columns(content: bytes) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    first_row = next(sheet.iter_rows(values_only=True), [])
    return [str(value) for value in first_row if value is not None]
